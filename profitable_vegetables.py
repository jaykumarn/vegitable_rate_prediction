"""
profitable_vegetables.py
========================
Reads  product_all.xlsx, keeps only vegetable rows (code_number 1001-1004 and
2001-2043), and, for every calendar month in the data, ranks each vegetable
on three equal-weight criteria:

    1. Price          – mean of (max_rate + min_rate) / 2  across all days in
                        the month  (Rs per quintal / per 100).
    2. Volume         – total quantity traded in the month  (quintals / hundreds).
    3. Per-Acre Prod  – total_volume ÷ number_of_trading_days.  This is the
                        best proxy for yield / per-acre productivity that the
                        data contains; it captures how much of each vegetable
                        moves on a single market day.

Scoring method  (Min-Max normalisation, equal weights)
------------------------------------------------------
Within every month the three raw values are independently scaled to [0, 1]
using  (x − min) / (max − min).   The composite Profit Score is the simple
average of the three normalised scores.   Because all normalisation windows
are the same month, every score is directly comparable across vegetables
within that month.

Outputs
-------
* Console   – top-10 most-profitable vegetables for every month, with a
              short explanation block printed once at the start.
* Excel     – "profitable_vegetables.xlsx"
                Sheet "Top 10 Per Month"  – the ranked top-10 table for each
                                            month, colour-coded by rank.
                Sheet "Full Scores"       – every vegetable × month combination
                                            with all raw and normalised columns.
                Sheet "Score Breakdown"   – pivot tables (one per metric) so a
                                            reader can visually compare any
                                            vegetable across months.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# 1.  LOAD  &  FILTER  –  keep only vegetables
# ──────────────────────────────────────────────
INPUT_FILE  = "product_all.xlsx"
OUTPUT_FILE = "profitable_vegetables.xlsx"

VEGETABLE_CODES = list(range(1001, 1005)) + list(range(2001, 2044))

df = pd.read_excel(
    INPUT_FILE,
    sheet_name=0,
    header=0,
    usecols=[
        "rate_date", "code_number", "product_name",
        "product_quantity", "product_max_rate", "product_min_rate",
    ],
)

df = df[df["code_number"].isin(VEGETABLE_CODES)].copy()

# ──────────────────────────────────────────────
# 2.  CLEAN  –  drop &nbsp; placeholders, parse
#              "Rs. 1700/-" strings → float
# ──────────────────────────────────────────────
df = df[df["product_quantity"].astype(str).str.strip() != "&nbsp;"].copy()


def _parse_rs(value: str) -> float:
    """'Rs. 1700/-'  →  1700.0"""
    cleaned = str(value).replace("Rs.", "").replace("/-", "").replace(",", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return np.nan


df["qty"]       = pd.to_numeric(df["product_quantity"], errors="coerce")
df["max_rate"]  = df["product_max_rate"].apply(_parse_rs)
df["min_rate"]  = df["product_min_rate"].apply(_parse_rs)
df["avg_rate"]  = (df["max_rate"] + df["min_rate"]) / 2.0

df.dropna(subset=["qty", "max_rate", "min_rate"], inplace=True)

df["rate_date"]  = pd.to_datetime(df["rate_date"])
df["year_month"] = df["rate_date"].dt.to_period("M")

# ──────────────────────────────────────────────
# 3.  MONTHLY AGGREGATES  per vegetable
# ──────────────────────────────────────────────
monthly = (
    df.groupby(["year_month", "code_number", "product_name"])
    .agg(
        avg_price    =("avg_rate",   "mean"),   # mean daily avg-price
        total_volume =("qty",        "sum"),    # total quintals / hundreds
        n_days       =("qty",        "count"),  # number of trading days
    )
    .reset_index()
)

# Per-Acre Production proxy  =  total volume  /  trading days
monthly["per_acre"] = monthly["total_volume"] / monthly["n_days"]

# ──────────────────────────────────────────────
# 4.  MIN-MAX NORMALISE  (per-month)  &  SCORE
# ──────────────────────────────────────────────
def _minmax(series: pd.Series) -> pd.Series:
    lo, hi = series.min(), series.max()
    return (series - lo) / (hi - lo) if hi != lo else pd.Series(0.0, index=series.index)


scored_frames = []

for _, month_df in monthly.groupby("year_month"):
    s = month_df.copy()
    s["norm_price"]    = _minmax(s["avg_price"]).values
    s["norm_volume"]   = _minmax(s["total_volume"]).values
    s["norm_per_acre"] = _minmax(s["per_acre"]).values

    # Equal weight  →  simple average of the three normalised scores
    s["profit_score"] = (
        s["norm_price"] + s["norm_volume"] + s["norm_per_acre"]
    ) / 3.0

    scored_frames.append(s)

scored = pd.concat(scored_frames, ignore_index=True)
scored["year_month_str"] = scored["year_month"].astype(str)

# ──────────────────────────────────────────────
# 5.  TOP-10 TABLE  per month
# ──────────────────────────────────────────────
top10_rows = []
for ym in sorted(scored["year_month"].unique()):
    subset = scored[scored["year_month"] == ym].nlargest(10, "profit_score").copy()
    subset["rank"] = range(1, len(subset) + 1)
    top10_rows.append(subset)

top10 = pd.concat(top10_rows, ignore_index=True)

# ──────────────────────────────────────────────
# 6.  CONSOLE OUTPUT
# ──────────────────────────────────────────────
EXPLAIN = """
╔══════════════════════════════════════════════════════════════════════════════╗
║              PROFITABLE VEGETABLES REPORT  –  HOW IT WORKS                ║
╠══════════════════════════════════════════════════════════════════════════════╣
║                                                                            ║
║  Source   product_all.xlsx   →  vegetable codes 1001-1004  &  2001-2043    ║
║  Period   June 2024  –  November 2025  (18 calendar months)                ║
║                                                                            ║
║  Three criteria, each given EQUAL weight (33.3 %)                          ║
║  ─────────────────────────────────────────────────                         ║
║  1. Price          Mean of (Max Rate + Min Rate)/2  across every           ║
║                    trading day in the month.                               ║
║  2. Volume         Total quantity (quintals/hundreds) traded               ║
║                    across the whole month.                                 ║
║  3. Per-Acre Prod  Total Volume ÷ Number of Trading Days.                  ║
║                    Proxy for daily yield / per-acre productivity.          ║
║                                                                            ║
║  Scoring                                                                   ║
║  ───────                                                                   ║
║  • Each metric is Min-Max normalised to [0, 1] inside its own month.       ║
║  • Profit Score  =  (norm_price + norm_volume + norm_per_acre) / 3         ║
║  • Rank 1 is the most profitable vegetable for that month.                 ║
║                                                                            ║
║  Output                                                                    ║
║  ──────                                                                    ║
║  • Console  – Top 10 for every month.                                      ║
║  • Excel    – "profitable_vegetables.xlsx"                                 ║
║       Sheet 1  "Top 10 Per Month"   – colour-coded ranked table.           ║
║       Sheet 2  "Full Scores"        – every veg × month with all columns.  ║
║       Sheet 3  "Score Breakdown"    – pivot tables per metric.             ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""
print(EXPLAIN)

DISPLAY_COLS = [
    "product_name", "avg_price", "total_volume", "per_acre",
    "norm_price", "norm_volume", "norm_per_acre", "profit_score",
]

for ym in sorted(scored["year_month"].unique()):
    subset = (
        scored[scored["year_month"] == ym]
        .nlargest(10, "profit_score")[DISPLAY_COLS]
    )
    print(f"\n{'─'*88}")
    print(f"  {ym}  –  Top 10 Most Profitable Vegetables")
    print(f"{'─'*88}")
    print(
        subset.to_string(
            index=False,
            float_format=lambda x: f"{x:.2f}",
        )
    )

# ──────────────────────────────────────────────
# 7.  WRITE EXCEL  (three sheets, styled)
# ──────────────────────────────────────────────

# ── colour palette ────────────────────────────
RANK_FILLS = {
    1:  PatternFill("solid", fgColor="006400"),   # dark green
    2:  PatternFill("solid", fgColor="228B22"),   # forest green
    3:  PatternFill("solid", fgColor="32CD32"),   # lime green
    4:  PatternFill("solid", fgColor="66CDAA"),   # medium aquamarine
    5:  PatternFill("solid", fgColor="7EC8E3"),   # light blue
    6:  PatternFill("solid", fgColor="87CEEB"),   # sky blue
    7:  PatternFill("solid", fgColor="ADD8E6"),   # light blue 2
    8:  PatternFill("solid", fgColor="FFD700"),   # gold
    9:  PatternFill("solid", fgColor="FFA07A"),   # light salmon
    10: PatternFill("solid", fgColor="FF6347"),   # tomato
}
WHITE_FONT = Font(name="Calibri", size=10, bold=False, color="FFFFFF")
BOLD_WHITE = Font(name="Calibri", size=10, bold=True,  color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F4F4F")   # dark slate grey
HEADER_FONT = Font(name="Calibri", size=10, bold=True,  color="FFFFFF")
MONTH_FILL  = PatternFill("solid", fgColor="1B3A4B")
MONTH_FONT  = Font(name="Calibri", size=11, bold=True,  color="FFFFFF")
THIN_BORDER = Border(
    left  =Side(style="thin", color="B0B0B0"),
    right =Side(style="thin", color="B0B0B0"),
    top   =Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# column headers that go into Sheet-1
TOP10_HEADERS = [
    "Rank",
    "Vegetable Name",
    "Avg Price (Rs)",
    "Total Volume",
    "Per-Acre Prod",
    "Norm Price",
    "Norm Volume",
    "Norm Per-Acre",
    "Profit Score",
]
TOP10_COL_WIDTHS = [6, 22, 14, 15, 14, 12, 12, 14, 13]


def _style_header_row(ws, row_idx: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_month_row(ws, row_idx: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font  = MONTH_FONT
        cell.fill  = MONTH_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER


# ── Sheet 1: "Top 10 Per Month" ──────────────
def write_top10_sheet(ws):
    n_cols = len(TOP10_HEADERS)
    # column widths
    for i, w in enumerate(TOP10_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # global header row  (row 1)
    for i, h in enumerate(TOP10_HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, n_cols)
    ws.row_dimensions[1].height = 22

    current_row = 2

    for ym in sorted(top10["year_month"].unique()):
        # ── month banner row ──
        month_label = f"  {ym.strftime('%B %Y')}  –  Top 10 Most Profitable Vegetables"
        ws.cell(row=current_row, column=1, value=month_label)
        _style_month_row(ws, current_row, n_cols)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # ── data rows (rank 1-10) ──
        month_data = (
            top10[top10["year_month"] == ym]
            .nlargest(10, "profit_score")
            .reset_index(drop=True)
        )

        for idx, row in month_data.iterrows():
            rank = idx + 1
            values = [
                rank,
                row["product_name"],
                round(row["avg_price"], 2),
                int(row["total_volume"]),
                round(row["per_acre"], 2),
                round(row["norm_price"], 4),
                round(row["norm_volume"], 4),
                round(row["norm_per_acre"], 4),
                round(row["profit_score"], 4),
            ]
            fill = RANK_FILLS.get(rank, PatternFill())
            font = BOLD_WHITE if rank <= 3 else WHITE_FONT

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.fill   = fill
                cell.font   = font
                cell.border = THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="center" if col_idx != 2 else "left",
                    vertical="center",
                )
                # numeric formats
                if col_idx in (3, 5):          # price / per-acre
                    cell.number_format = "#,##0.00"
                elif col_idx == 4:             # volume
                    cell.number_format = "#,##0"
                elif col_idx in (6, 7, 8, 9):  # normalised / score
                    cell.number_format = "0.0000"

            ws.row_dimensions[current_row].height = 18
            current_row += 1

        current_row += 1          # blank row between months

    # freeze panes: keep header visible while scrolling
    ws.freeze_panes = ws.cell(row=2, column=1)


# ── Sheet 2: "Full Scores" ───────────────────
def write_full_sheet(ws):
    cols = [
        "year_month_str", "code_number", "product_name",
        "avg_price", "total_volume", "n_days", "per_acre",
        "norm_price", "norm_volume", "norm_per_acre", "profit_score",
    ]
    headers = [
        "Month", "Code", "Vegetable Name",
        "Avg Price (Rs)", "Total Volume", "Trading Days", "Per-Acre Prod",
        "Norm Price", "Norm Volume", "Norm Per-Acre", "Profit Score",
    ]
    widths = [12, 8, 22, 14, 14, 13, 14, 12, 12, 14, 13]

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # header
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    # data – sorted by month then profit_score descending
    export = scored.sort_values(["year_month_str", "profit_score"], ascending=[True, False])

    for r_idx, (_, row) in enumerate(export.iterrows(), start=2):
        vals = [row[c] for c in cols]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.font   = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="center" if c_idx != 3 else "left")
            if c_idx in (4, 7):
                cell.number_format = "#,##0.00"
            elif c_idx in (5, 6):
                cell.number_format = "#,##0"
            elif c_idx in (8, 9, 10, 11):
                cell.number_format = "0.0000"

    ws.freeze_panes = ws.cell(row=2, column=1)


# ── Sheet 3: "Score Breakdown" (pivot tables) ──
def write_breakdown_sheet(ws):
    pivot_specs = [
        ("Avg Price  (Rs / quintal)",   "avg_price"),
        ("Total Volume  (quintals)",    "total_volume"),
        ("Per-Acre Production",         "per_acre"),
        ("Profit Score",                "profit_score"),
    ]

    months       = sorted(scored["year_month_str"].unique())
    veg_names    = sorted(scored["product_name"].unique())
    n_months     = len(months)
    n_vegs       = len(veg_names)
    n_cols       = n_months + 1          # vegetable-name col + one per month

    PIVOT_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
    PIVOT_HEADER_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    VEG_COL_FILL      = PatternFill("solid", fgColor="D6E4F0")
    VEG_COL_FONT      = Font(name="Calibri", size=9, bold=True, color="1B3A4B")
    TITLE_FONT        = Font(name="Calibri", size=12, bold=True, color="1B3A4B")

    current_row = 1

    for title, metric in pivot_specs:
        # ── title row ──
        ws.cell(row=current_row, column=1, value=title)
        ws.cell(row=current_row, column=1).font = TITLE_FONT
        current_row += 1

        # ── header row  ["Vegetable", "2024-06", "2024-07", ...] ──
        ws.cell(row=current_row, column=1, value="Vegetable")
        ws.cell(row=current_row, column=1).font  = PIVOT_HEADER_FONT
        ws.cell(row=current_row, column=1).fill  = PIVOT_HEADER_FILL
        ws.cell(row=current_row, column=1).border = THIN_BORDER
        ws.column_dimensions["A"].width = 22

        for m_idx, m_label in enumerate(months, 2):
            cell = ws.cell(row=current_row, column=m_idx, value=m_label)
            cell.font   = PIVOT_HEADER_FONT
            cell.fill   = PIVOT_HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(m_idx)].width = 13

        current_row += 1

        # ── build a quick lookup  (veg_name, month_str) → value ──
        lookup = {}
        for _, row in scored.iterrows():
            lookup[(row["product_name"], row["year_month_str"])] = row[metric]

        # ── data rows ──
        for veg in veg_names:
            # vegetable-name cell
            cell = ws.cell(row=current_row, column=1, value=veg)
            cell.font   = VEG_COL_FONT
            cell.fill   = VEG_COL_FILL
            cell.border = THIN_BORDER

            for m_idx, m_label in enumerate(months, 2):
                val  = lookup.get((veg, m_label), "")
                cell = ws.cell(row=current_row, column=m_idx, value=val)
                cell.border = THIN_BORDER
                cell.font   = Font(name="Calibri", size=9)
                cell.alignment = Alignment(horizontal="center")

                if val != "":
                    if metric in ("avg_price", "per_acre"):
                        cell.number_format = "#,##0.00"
                    elif metric == "total_volume":
                        cell.number_format = "#,##0"
                    else:                         # profit_score
                        cell.number_format = "0.0000"

            current_row += 1

        current_row += 2        # two blank rows between pivot tables


# ── assemble workbook ─────────────────────────
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    # write empty DataFrames just to create the sheets with the right names
    pd.DataFrame().to_excel(writer, sheet_name="Top 10 Per Month", index=False)
    pd.DataFrame().to_excel(writer, sheet_name="Full Scores",      index=False)
    pd.DataFrame().to_excel(writer, sheet_name="Score Breakdown",  index=False)

wb = load_workbook(OUTPUT_FILE)

# remove auto-created empty sheets if any duplicates exist
for name in list(wb.sheetnames):
    if name not in ("Top 10 Per Month", "Full Scores", "Score Breakdown"):
        del wb[name]

write_top10_sheet(wb["Top 10 Per Month"])
write_full_sheet(wb["Full Scores"])
write_breakdown_sheet(wb["Score Breakdown"])

wb.save(OUTPUT_FILE)
print(f"\n✅  Excel report saved  →  {OUTPUT_FILE}")
print(f"     Sheet 1 : Top 10 Per Month   –  colour-coded ranked tables")
print(f"     Sheet 2 : Full Scores        –  every vegetable × month")
print(f"     Sheet 3 : Score Breakdown    –  pivot tables for each metric")
